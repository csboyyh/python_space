import os
import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin
from openpyxl import load_workbook, Workbook
from time import sleep
import random
from fake_useragent import UserAgent

EXCEL_FILE = "extracted_info.xlsx"


def get_html_content(url):
    ua = UserAgent()
    headers = {"User-Agent": ua.random}
    delay = random.randint(10, 20)
    sleep(delay)
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        html_content = response.text
        return html_content
    else:
        print(f"Request failed for URL: {url}. Retrying after 30 minutes.")
        sleep(1800)
        return get_html_content(url)


def extract_product_details(product_link):
    print(f"Extracting details from {product_link}")
    html_content = get_html_content(product_link)
    soup = BeautifulSoup(html_content, "html.parser")
    details = {}

    specs_list = soup.find("div", id="specs-list")

    if specs_list:
        rows = specs_list.find_all("tr")
        for row in rows:
            cells = row.find_all("td")
            if len(cells) == 2:
                detail_name = cells[0].text.strip()
                detail_value = cells[1].text.strip()
                # Check if the detail_name already exists in the dictionary
                if detail_name in details:
                    # If it does, append the new value
                    details[detail_name] += "; " + detail_value
                else:
                    # Otherwise, just add the detail_name and detail_value
                    details[detail_name] = detail_value

    return details


def extract_product_info(brand, product_list, full_link, workbook, soup):
    sheet = workbook.active
    for product in product_list:
        product_info = product.find_all("a")
        for info in product_info:
            product_name = info.text.strip()
            product_link = urljoin(full_link, info["href"])

            if check_product_link_exists(product_link, sheet):
                print(f"Skipping {product_link}. Already exists in the Excel.")
                continue

            product_details = extract_product_details(product_link)
            print(f"Extracted details for {product_name}: {product_details}")
            row_data = [brand, product_name, product_link]
            row_data.extend([f"{name}: {value}" for name, value in product_details.items()])
            sheet.append(row_data)
            workbook.save(EXCEL_FILE)

    workbook.save(EXCEL_FILE)


def check_product_link_exists(product_link, sheet):
    for row in sheet.iter_rows(min_row=2, values_only=True, max_col=3):
        if row[2] == product_link:
            return True
    return False

def check_brand_exists(brand, sheet):
    for row in sheet.iter_rows(min_row=2, values_only=True, max_col=3):
        if row[0] == brand:
            return True
    return False
    
def extract_info_to_excel(url):
    print(f"Extracting information from {url}")
    html_content = get_html_content(url)
    soup = BeautifulSoup(html_content, "html.parser")
    tds = soup.find_all("td")
    
    
    if os.path.exists(EXCEL_FILE):
        workbook = load_workbook(EXCEL_FILE)
        sheet = workbook.active
    else:
        workbook = Workbook()
        workbook.active.append(["Brand", "Product Name", "Product Link"])
        sheet = workbook.active		
		
    for td in tds:
    	href = td.a["href"]
    	full_link = urljoin(url, href)
    	brand = td.a.text.strip()
    	if check_brand_exists(brand, sheet):
    		print(f"Skipping {brand}. Already exists in the Excel.")
    		continue
    	print(f"Processing brand: {brand}")
    	product_list = extract_product_list(full_link, soup)
    	extract_product_info(brand, product_list, full_link, workbook, soup)

    workbook.save(EXCEL_FILE)
    print("Extraction completed. Results saved to extracted_info.xlsx")


def extract_product_list(url, soup):
    print(f"Extracting product list from {url}")
    html_content = get_html_content(url)
    soup = BeautifulSoup(html_content, "html.parser")
    product_list = soup.find_all("div", class_="makers")
    pagination = soup.find("div", class_="nav-pages")
    if pagination:
        pages = pagination.find_all("a")
        for page in pages:
            page_url = urljoin(url, page["href"])
            print(f"Extracting product list from page: {page_url}")
            page_html_content = get_html_content(page_url)
            page_soup = BeautifulSoup(page_html_content, "html.parser")
            page_product_list = page_soup.find_all("div", class_="makers")
            product_list.extend(page_product_list)

    return product_list



if __name__ == "__main__":
    # 调用函数并提供目标 URL
    url = "https://www.gsmarena.com/makers.php3"
    extract_info_to_excel(url)